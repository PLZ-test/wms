# orders/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.db.models import Count, Q, F # [수정] F객체 추가
from datetime import datetime, date, timedelta
from django.db.models.functions import TruncDate
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
import openpyxl
import json

from management.models import Shipper, Product, SalesChannel
from stock.models import StockMovement # [추가] StockMovement 모델 import
from .models import Order, OrderItem
from .forms import OrderUpdateForm

# ... (order_manage_view, order_list_success_view 등 다른 뷰는 그대로 유지) ...
@login_required
def order_manage_view(request):
    """
    일자별 주문 관리 대시보드 뷰
    """
    from django.utils import timezone
    
    date_str = request.GET.get('date')
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    
    # 선택된 날짜의 00:00:00부터 23:59:59까지 (현재 시간대 기준)
    start_datetime = timezone.make_aware(datetime.combine(selected_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(selected_date, datetime.max.time()))
    
    # order_date가 해당 날짜 범위에 속하는 주문 필터링
    orders_for_date = Order.objects.filter(
        order_date__gte=start_datetime,
        order_date__lte=end_datetime
    )
    
    success_count = orders_for_date.exclude(order_status='ERROR').count()
    error_count = orders_for_date.filter(order_status='ERROR').count()
    
    # [삭제] 세션의 temp_errors 카운트 제거 (모든 오류가 DB에 저장됨)
    
    total_count = success_count + error_count
    
    daily_stats = {
        'total_count': total_count,
        'success_count': success_count,
        'error_count': error_count,
        'success_rate': (success_count / total_count * 100) if total_count > 0 else 0,
        'error_rate': (error_count / total_count * 100) if total_count > 0 else 0,
    }
    context = {
        'page_title': '주문 관리',
        'active_menu': 'orders',
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'daily_stats': daily_stats,
    }
    return render(request, 'orders/order_manage.html', context)

@login_required
def order_list_success_view(request, date_str):
    """
    선택한 날짜의 '성공' 주문 목록을 보여주는 뷰
    """
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    print(f"DEBUG: Success View Called using TEMPLATE: orders/order_list_result_new.html")
    orders = Order.objects.filter(order_date__date=target_date).exclude(order_status='ERROR').prefetch_related('items__product')
    
    orders_json = []
    for order in orders:
        items = [{'product_name': item.product.name, 'quantity': item.quantity} for item in order.items.all()]
        orders_json.append({'id': order.id, 'items': items})
        
    context = {
        'page_title': f'{date_str} 성공 주문 목록',
        'orders': orders,
        'list_type': 'success',
        'orders_json': orders_json,
    }
    return render(request, 'orders/order_list_final.html', context)

@login_required
def order_list_error_view(request, date_str):
    """
    선택한 날짜의 '오류' 주문 목록을 보여주는 뷰
    """
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    print(f"DEBUG: Error View Called using TEMPLATE: orders/order_list_result_new.html")
    processed_errors = []

    # [수정] DB에서만 오류 조회 (세션 로직 제거)
    db_error_orders = Order.objects.filter(order_date__date=target_date, order_status='ERROR')
    for order in db_error_orders:
        error_details = {}
        try:
            error_details = json.loads(order.error_message)
        except (json.JSONDecodeError, TypeError):
            pass
        
        original_data = error_details.get('original_data', {})
        processed_errors.append({
            'is_db': True,
            'unique_id': f"db-{order.id}",
            'id': order.id,
            'order_no': order.order_no,
            'recipient_name': order.recipient_name,
            'shipper_name': order.shipper.name if order.shipper else original_data.get('shipper_name', ''),
            'product_identifier': original_data.get('product_identifier', ''),
            'channel_name': order.channel.name if order.channel else original_data.get('channel_name', ''),
            'quantity': original_data.get('quantity', ''),
            'recipient_phone': order.recipient_phone,
            'address': order.address,
            'error_message': error_details.get('error_message', order.error_message or '알 수 없는 오류'),
            'error_fields': error_details.get('error_fields', []),
        })

    # [삭제] 세션 오류 조회 로직 제거

    context = {
        'page_title': f'{date_str} 오류 주문 목록', 
        'orders': processed_errors,
        'list_type': 'error',
    }
    return render(request, 'orders/order_list_final.html', context)

@login_required
def order_update_view(request, order_pk):
    """
    오류 주문 상세 수정 페이지 뷰
    """
    order = get_object_or_404(Order, pk=order_pk, order_status='ERROR')
    if request.method == 'POST':
        form = OrderUpdateForm(request.POST, instance=order)
        if form.is_valid():
            updated_order = form.save(commit=False)
            updated_order.error_message = ""
            updated_order.order_status = 'PENDING'
            updated_order.save()
            messages.success(request, f"주문({order.order_no})이 성공적으로 수정되었습니다.")
            date_str = order.order_date.strftime('%Y-%m-%d')
            return redirect('orders:list_error', date_str=date_str)
    else:
        try:
            error_details = json.loads(order.error_message)
            initial_data = error_details.get('original_data', {})
            initial_data.update({ 'recipient_name': order.recipient_name, 'recipient_phone': order.recipient_phone, 'address': order.address })
            form = OrderUpdateForm(initial=initial_data)
        except (json.JSONDecodeError, TypeError):
            form = OrderUpdateForm(instance=order)
            
    context = { 'page_title': '오류 주문 수정', 'form': form, 'order': order }
    return render(request, 'orders/order_update_form.html', context)


@login_required
@require_POST
@transaction.atomic # [추가] 데이터베이스 트랜잭션 적용
def order_invoice_view(request):
    """
    송장 출력 및 자동 출고 처리 뷰
    """
    order_ids_str = request.POST.get('order_ids', '')
    if not order_ids_str:
        return HttpResponse("출력할 주문이 선택되지 않았습니다.", status=400)
    
    order_ids = [int(id) for id in order_ids_str.split(',')]
    # [수정] 출고 처리가 아직 안된 주문만 대상으로 함
    orders = Order.objects.filter(
        id__in=order_ids, 
        order_status__in=['PENDING', 'PROCESSING']
    ).select_related('shipper__center').prefetch_related('items__product')
    
    if not orders:
        return HttpResponse("출력할 주문이 없거나 이미 처리된 주문입니다.", status=404)
        
    # --- [신규] 자동 출고 처리 로직 ---
    for order in orders:
        for item in order.items.all():
            product = item.product
            # 1. 재고 수량 확인
            if product.quantity < item.quantity:
                messages.error(request, f"재고 부족: '{product.name}'의 출고를 처리할 수 없습니다. (현재 재고: {product.quantity})")
                # 트랜잭션을 롤백하고 함수를 종료
                transaction.set_rollback(True)
                # 이전 페이지로 리디렉션
                return redirect(request.META.get('HTTP_REFERER', 'orders:manage'))

            # 2. 재고 수량 차감
            product.quantity = F('quantity') - item.quantity
            product.save()

            # 3. 출고 기록(StockMovement) 생성
            StockMovement.objects.create(
                product=product,
                movement_type='OUT',
                quantity=item.quantity,
                memo=f'주문 출고 ({order.order_no})'
            )
    # ------------------------------------

    # 송장 출력 시 주문 상태를 '출고완료'로 변경
    orders.update(order_status='SHIPPED')
    
    context = {'orders': orders}
    return render(request, 'orders/invoice_template.html', context)

# ... (이하 API 뷰들은 그대로 유지) ...
@login_required
@require_POST
@transaction.atomic
def process_orders_api(request):
    """
    업로드된 엑셀 파일을 처리하여 주문을 생성하는 API
    """
    excel_file = request.FILES.get('excel_file')
    handle_duplicates = request.POST.get('handle_duplicates')
    if not excel_file:
        return JsonResponse({'status': 'error', 'message': '엑셀 파일이 없습니다.'}, status=400)

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active
    
    success_orders, error_orders = [], []
    processed_in_this_file = set()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        order_data = {
            'order_no': str(row[0]).strip() if row[0] else None, 'shipper_name': str(row[1]).strip() if row[1] else None,
            'channel_name': str(row[2]).strip() if row[2] else None, 'recipient_name': str(row[3]).strip() if row[3] else None,
            'recipient_phone': str(row[4]).strip() if row[4] else '', 'address': str(row[5]).strip() if row[5] else '',
            'product_identifier': str(row[6]).strip() if row[6] else None, 'quantity': int(row[7]) if row[7] and str(row[7]).isdigit() else 0,
        }
        
        errors, error_fields = [], []
        shipper = None

        try:
            if not order_data['shipper_name']: errors.append("화주사 정보 누락"); error_fields.append('shipper_name')
            else:
                try: shipper = Shipper.objects.get(name=order_data['shipper_name'])
                except Shipper.DoesNotExist: errors.append("미등록 화주사"); error_fields.append('shipper_name')

            if not order_data['product_identifier']: errors.append("상품 정보 누락"); error_fields.append('product_identifier')
            elif shipper and not Product.objects.filter(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper).exists():
                errors.append("미등록 상품"); error_fields.append('product_identifier')
            
            if not order_data['quantity'] or order_data['quantity'] <= 0: errors.append("수량 오류"); error_fields.append('quantity')
            if not order_data['channel_name']: errors.append("판매채널 누락"); error_fields.append('channel_name')
            if errors: raise ValueError(", ".join(sorted(list(set(errors)))))

            if handle_duplicates != 'yes':
                if Order.objects.filter(shipper=shipper, recipient_name=order_data['recipient_name'], address=order_data['address'], recipient_phone=order_data['recipient_phone']).exists(): continue
                file_dup_key = (order_data['shipper_name'], order_data['recipient_name'], order_data['address'], order_data['recipient_phone'])
                if file_dup_key in processed_in_this_file: continue
                processed_in_this_file.add(file_dup_key)

            channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
            product = Product.objects.get(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper)
            
            order = Order.objects.create(
                order_no=order_data['order_no'], shipper=shipper, channel=channel,
                recipient_name=order_data['recipient_name'], recipient_phone=order_data['recipient_phone'],
                address=order_data['address'], order_date=datetime.now(), order_status='PENDING'
            )
            OrderItem.objects.create(order=order, product=product, quantity=order_data['quantity'])
            success_orders.append(order)

        except Exception as e:
            # [수정] 오류를 세션 대신 DB에 저장
            # channel 정보 처리
            channel = None
            if order_data.get('channel_name'):
                channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
            
            # DB에 오류 주문 생성
            error_order = Order.objects.create(
                order_no=order_data.get('order_no'),
                shipper=shipper,  # shipper가 없는 경우 None
                channel=channel,  # channel이 없는 경우 None
                recipient_name=order_data.get('recipient_name', ''),
                recipient_phone=order_data.get('recipient_phone', ''),
                address=order_data.get('address', ''),
                order_date=datetime.now(),
                order_status='ERROR',
                error_message=json.dumps({
                    'error_message': str(e),
                    'error_fields': sorted(list(set(error_fields))),
                    'original_data': order_data
                }, ensure_ascii=False)
            )
            error_orders.append(error_order)

    messages.success(request, f"엑셀 처리가 완료되었습니다. 성공: {len(success_orders)}건, 실패: {len(error_orders)}건")
    # [삭제] 세션에 오류 저장하는 로직 제거
    # request.session['temp_errors'] = error_orders
    today_str = date.today().strftime('%Y-%m-%d')
    return JsonResponse({'status': 'success', 'redirect_url': reverse('orders:manage') + f'?date={today_str}'})

@login_required
@require_POST
@transaction.atomic
def batch_retry_error_api(request):
    """
    오류 목록 페이지에서 인라인 수정 후 일괄 재시도하는 API
    """
    data = json.loads(request.body)
    
    # [수정] 클라이언트가 { updates: [...], ... } 형태로 보낼 수도 있고, 바로 [...] 리스트로 보낼 수도 있음
    if isinstance(data, dict):
        all_items_data = data.get('updates', [])
    else:
        all_items_data = data
        
    results = []
    
    for item_data in all_items_data:
        order_data = item_data.get('data', {})
        unique_id = item_data.get('unique_id')
        
        # [수정] unique_id가 없거나 형식이 잘못된 경우 방어
        if not unique_id:
             results.append({'unique_id': 'unknown', 'status': 'error', 'error_message': 'ID 누락'})
             continue

        errors, error_fields = [], []; shipper = None
        try:
            if not order_data.get('shipper_name'): errors.append("화주사 정보 누락"); error_fields.append('shipper_name')
            else:
                try: shipper = Shipper.objects.get(name=order_data['shipper_name'])
                except Shipper.DoesNotExist: errors.append("미등록 화주사"); error_fields.append('shipper_name')
            
            if not order_data.get('product_identifier'): errors.append("상품 정보 누락"); error_fields.append('product_identifier')
            elif shipper and not Product.objects.filter(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper).exists():
                errors.append("미등록 상품"); error_fields.append('product_identifier')

            if not order_data.get('quantity') or int(order_data.get('quantity', 0)) <= 0: errors.append("수량 오류"); error_fields.append('quantity')
            if errors: raise ValueError(", ".join(sorted(list(set(errors)))))
            
            channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
            product = Product.objects.get(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper)
            
            new_order = Order.objects.create(
                order_no=order_data.get('order_no') or None, shipper=shipper, channel=channel,
                recipient_name=order_data['recipient_name'], recipient_phone=order_data['recipient_phone'],
                address=order_data['address'], order_date=datetime.now(), order_status='PENDING'
            )
            OrderItem.objects.create(order=new_order, product=product, quantity=order_data['quantity'])
            
            # 성공 시 원본 오류 삭제
            try:
                id_type, id_value = unique_id.split('-')
                if id_type == 'db': 
                    Order.objects.filter(id=id_value).delete()
            except ValueError:
                pass # ID 형식이 다르더라도 새로운 주문은 생성되었으므로 진행
            
            results.append({'unique_id': unique_id, 'status': 'success'})

        except Exception as e:
            # [수정] 실패 시 DB의 오류 Order를 업데이트
            error_message = str(e)
            results.append({ 'unique_id': unique_id, 'status': 'error', 'error_message': error_message, 'error_fields': sorted(list(set(error_fields))) })
            
            # DB 오류 업데이트
            try:
                id_type, id_value = unique_id.split('-')
                if id_type == 'db':
                    error_order = Order.objects.get(id=id_value, order_status='ERROR')
                    # 오류 정보 업데이트
                    error_order.error_message = json.dumps({
                        'error_message': error_message,
                        'error_fields': sorted(list(set(error_fields))),
                        'original_data': order_data
                    }, ensure_ascii=False)
                    # 날짜를 현재 날짜로 업데이트
                    error_order.order_date = datetime.now()
                    # 수정된 수취인 정보 업데이트
                    error_order.recipient_name = order_data.get('recipient_name', '')
                    error_order.recipient_phone = order_data.get('recipient_phone', '')
                    error_order.address = order_data.get('address', '')
                    error_order.save()
            except (Order.DoesNotExist, ValueError):
                pass  # 이미 삭제된 경우거나 ID 형식이 잘못된 경우 무시
            
    return JsonResponse({'results': results})

@login_required
def product_autocomplete_api(request):
    """
    오류 수정 시 상품명/바코드 자동완성을 위한 API
    """
    shipper_name = request.GET.get('shipper_name'); term = request.GET.get('term')
    if not shipper_name or not term: return JsonResponse([], safe=False)
    products = Product.objects.filter(Q(name__icontains=term) | Q(barcode__icontains=term), shipper__name=shipper_name).values_list('name', flat=True)[:10]
    return JsonResponse(list(products), safe=False)

@login_required
def order_chart_data_api(request):
    """
    홈 대시보드의 일별 주문 추이 차트 데이터를 제공하는 API
    """
    start_str = request.GET.get('start'); end_str = request.GET.get('end')
    try: start_date, end_date = datetime.strptime(start_str, '%Y-%m-%d').date(), datetime.strptime(end_str, '%Y-%m-%d').date()
    except (ValueError, TypeError): end_date, start_date = date.today(), date.today() - timedelta(days=6)
    
    labels = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end_date - start_date).days + 1)]
    status_map = { 'PENDING': '주문접수', 'PROCESSING': '처리중', 'ERROR': '오류' }
    
    orders = Order.objects.filter(order_date__date__range=[start_date, end_date], order_status__in=status_map.keys()).annotate(date=TruncDate('order_date')).values('date', 'order_status').annotate(count=Count('id')).order_by('date')
    
    daily_counts = {label: {status: 0 for status in status_map} for label in labels}
    for order in orders:
        date_str = order['date'].strftime('%Y-%m-%d')
        if date_str in daily_counts: daily_counts[date_str][order['order_status']] = order['count']
            
    colors = { 'PENDING': 'rgba(54, 162, 235, 0.7)', 'PROCESSING': 'rgba(255, 159, 64, 0.7)', 'ERROR': 'rgba(255, 99, 132, 0.7)' }
    datasets = []
    for code, name in status_map.items():
        data = [daily_counts[label][code] for label in labels]
        if any(d > 0 for d in data):
            datasets.append({ 'label': name, 'data': data, 'borderColor': colors.get(code), 'backgroundColor': colors.get(code), 'fill': False, 'tension': 0.1 })
            
    return JsonResponse({'labels': labels, 'datasets': datasets})

@login_required
def channel_order_chart_data_api(request):
    """
    주문 관리 페이지의 채널별 주문량 원형 차트 데이터를 제공하는 API
    """
    date_str = request.GET.get('date')
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    
    orders_for_date = Order.objects.filter(order_date__date=target_date)
    channel_counts = SalesChannel.objects.filter(order__in=orders_for_date).annotate(order_count=Count('order')).values('name', 'order_count')
    
    labels = [data['name'] for data in channel_counts]
    data = [data['order_count'] for data in channel_counts]
    return JsonResponse({'labels': labels, 'data': data})


# --- 주문 취소 기능 (재고 복구 포함) ---

@login_required
@require_POST
@transaction.atomic
def order_cancel_view(request, order_pk):
    """
    주문 취소 및 재고 복구
    """
    order = get_object_or_404(Order, pk=order_pk)
    
    # 이미 취소된 주문은 처리 불가
    if order.order_status == 'CANCELED':
        messages.warning(request, "이미 취소된 주문입니다.")
        return redirect(request.META.get('HTTP_REFERER', 'orders:manage'))
    
    # 배송 완료된 주문은 취소 불가
    if order.order_status == 'DELIVERED':
        messages.error(request, "배송 완료된 주문은 취소할 수 없습니다. 반품 처리로 진행해주세요.")
        return redirect(request.META.get('HTTP_REFERER', 'orders:manage'))
    
    # 출고된 주문이면 재고 복구
    if order.order_status == 'SHIPPED':
        for item in order.items.all():
            product = item.product
            
            # 재고 복구 (F() 객체 사용으로 동시성 제어)
            product.quantity = F('quantity') + item.quantity
            product.save()
            
            # 재고 복구 후 실제 값을 다시 로드
            product.refresh_from_db()
            
            # 입고 기록 생성
            StockMovement.objects.create(
                product=product,
                movement_type='IN',
                quantity=item.quantity,
                memo=f'주문 취소로 인한 재고 복구 (주문번호: {order.order_no})'
            )
        
        messages.success(request, f"주문({order.order_no})이 취소되었으며, 재고가 복구되었습니다.")
    else:
        # 출고 전 주문은 재고 복구 없이 취소
        messages.success(request, f"주문({order.order_no})이 취소되었습니다.")
    
    # 주문 상태를 취소로 변경
    order.order_status = 'CANCELED'
    order.save()
    
    return redirect(request.META.get('HTTP_REFERER', 'orders:manage'))


# --- 주문 엑셀 내보내기 ---

@login_required
def order_export_excel_view(request):
    """
    주문 목록 엑셀 다운로드
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    date_str = request.GET.get('date')
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    
    # 해당 날짜의 주문 조회
    orders = Order.objects.filter(
        order_date__date=selected_date
    ).select_related('shipper', 'channel').prefetch_related('items__product').order_by('id')
    
    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"주문목록_{selected_date.strftime('%Y%m%d')}"
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 작성
    headers = [
        '주문번호', '주문일시', '화주사', '판매채널', '주문상태',
        '수취인', '연락처', '주소', '상품명', '바코드', '수량', '배송메모'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 데이터 작성
    row_num = 2
    for order in orders:
        # 주문 상태에 따른 색상
        if order.order_status == 'CANCELED':
            row_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        elif order.order_status == 'ERROR':
            row_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        elif order.order_status == 'DELIVERED':
            row_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
        else:
            row_fill = None
        
        for item in order.items.all():
            ws.cell(row=row_num, column=1, value=order.order_no)
            ws.cell(row=row_num, column=2, value=order.order_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=3, value=order.shipper.name if order.shipper else '')
            ws.cell(row=row_num, column=4, value=order.channel.name if order.channel else '')
            ws.cell(row=row_num, column=5, value=order.get_order_status_display())
            ws.cell(row=row_num, column=6, value=order.recipient_name)
            ws.cell(row=row_num, column=7, value=order.recipient_phone)
            ws.cell(row=row_num, column=8, value=order.address)
            ws.cell(row=row_num, column=9, value=item.product.name)
            ws.cell(row=row_num, column=10, value=item.product.barcode)
            ws.cell(row=row_num, column=11, value=item.quantity)
            ws.cell(row=row_num, column=12, value=order.delivery_memo)
            
            # 테두리 및 색상 적용
            for col_num in range(1, 13):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill
            
            row_num += 1
    
    # 주문이 없는 경우
    if row_num == 2:
        ws.cell(row=2, column=1, value="데이터가 없습니다.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        cell = ws.cell(row=2, column=1)
        cell.alignment = Alignment(horizontal="center")
    
    # 열 너비 자동 조정
    column_widths = {
        'A': 15,  # 주문번호
        'B': 17,  # 주문일시
        'C': 15,  # 화주사
        'D': 15,  # 판매채널
        'E': 12,  # 주문상태
        'F': 12,  # 수취인
        'G': 15,  # 연락처
        'H': 40,  # 주소
        'I': 30,  # 상품명
        'J': 15,  # 바코드
        'K': 8,   # 수량
        'L': 30,  # 배송메모
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # HTTP 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'orders_{selected_date.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


# --- 송장 출력 기능 ---

@login_required
def print_invoice(request, order_pk):
    """
    개별 주문 송장 출력
    """
    order = get_object_or_404(Order, pk=order_pk)
    context = {
        'orders': [order],
        'title': f'송장 - {order.order_no}'
    }
    return render(request, 'orders/invoice.html', context)

@login_required
def print_invoices_batch(request):
    """
    선택한 주문 일괄 송장 출력
    """
    order_ids = request.GET.get('ids', '').split(',')
    orders = Order.objects.filter(id__in=order_ids)
    
    if not orders.exists():
        return HttpResponse("선택된 주문이 없습니다.", status=404)
        
    context = {
        'orders': orders,
        'title': '일괄 송장 출력'
    }
    return render(request, 'orders/invoice.html', context)

@login_required
def download_sample_excel_view(request):
    """
    주문 업로드용 예시 엑셀 파일 다운로드
    """
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주문업로드_예시"

    headers = ['주문번호', '화주사', '판매채널', '수취인', '연락처', '주소', '상품명', '수량', '배송메모']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    data = [
        ['ORD-20240108-01', '테스트화주', '스마트스토어', '김철수', '010-1234-5678', '서울시 강남구 역삼동 123-45', '테스트상품A', 2, '안전배송'],
        ['ORD-20240108-02', '테스트화주', '쿠팡', '이영희', '010-9876-5432', '경기도 성남시 분당구 판교동 55', '테스트상품B', 1, '부재시 문앞'],
        ['ORD-20240108-03', '', '11번가', '박민수', '010-1111-2222', '부산시 해운대구 우동 77', '테스트상품A', 3, '빠른배송'],
        ['ORD-20240108-04', '테스트화주', 'G마켓', '최지우', '010-3333-4444', '대구시 수성구 범어동 88', '없는상품X', 1, ''],
        ['ORD-20240108-05', '테스트화주', '자사몰', '정우성', '010-5555-6666', '광주시 서구 치평동 99', '테스트상품B', 0, ''],
    ]

    for row in data:
        ws.append(row)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sample_orders.xlsx'
    wb.save(response)
    return response

@login_required
@require_POST
def delete_error_item_api(request):
    """
    오류 주문 항목 삭제 API (DB 또는 세션)
    """
    try:
        data = json.loads(request.body)
        unique_id = data.get('unique_id')
        
        if not unique_id:
            return JsonResponse({'status': 'error', 'message': 'ID 누락'}, status=400)
            
        id_type, id_value = unique_id.split('-')
        
        if id_type == 'db':
            # DB 삭제
            Order.objects.filter(id=id_value, order_status='ERROR').delete()
        elif id_type == 'session':
            # 세션 삭제 (더 이상 사용하지 않음, 하위 호환성 유지)
            pass
            
        return JsonResponse({'status': 'success'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# --- [신규] 쇼핑몰 API 주문 수집 ---

@login_required
@require_POST
def collect_orders_api(request):
    """
    쇼핑몰 API에서 수동으로 주문을 수집하는 API
    
    POST 데이터:
    - shipper_id: 특정 화주사 ID (선택)
    - channel_type: 특정 쇼핑몰 타입 (선택)
    - 둘 다 없으면 모든 활성화된 API에서 수집
    """
    from orders.services import OrderCollectorService
    
    try:
        data = json.loads(request.body)
        shipper_id = data.get('shipper_id')
        channel_type = data.get('channel_type')
        
        if shipper_id:
            # 특정 화주사의 주문 수집
            result = OrderCollectorService.collect_orders_for_shipper(
                shipper_id=shipper_id,
                channel_type=channel_type
            )
        else:
            # 모든 활성화된 API에서 수집
            result = OrderCollectorService.collect_all_active_orders()
        
        if result['status'] == 'success':
            messages.success(request, result['message'])
        else:
            messages.error(request, result['message'])
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'주문 수집 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


# --- [신규] 오류 주문 전체 취소 ---

@login_required
@require_POST
def cancel_all_errors_api(request):
    """
    특정 날짜의 모든 오류 주문을 취소하는 API
    
    POST 데이터:
    - date_str: 날짜 문자열 (YYYY-MM-DD)
    """
    try:
        from django.utils import timezone
        import json
        
        data = json.loads(request.body)
        date_str = data.get('date_str')
        
        if not date_str:
            return JsonResponse({
                'status': 'error',
                'message': '날짜가 지정되지 않았습니다.'
            }, status=400)
        
        # 날짜 파싱
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # 해당 날짜의 모든 ERROR 상태 주문 조회
        # 날짜 필터를 timezone aware하게 처리
        start_datetime = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(target_date, datetime.max.time()))
        
        error_orders = Order.objects.filter(
            order_date__gte=start_datetime,
            order_date__lte=end_datetime,
            order_status='ERROR'
        )
        
        count = error_orders.count()
        
        if count == 0:
            return JsonResponse({
                'status': 'info',
                'message': '취소할 오류 주문이 없습니다.'
            })
        
        # 모든 오류 주문 삭제
        error_orders.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': f'{count}건의 오류 주문을 취소했습니다.'
        })
        
    except ValueError:
        return JsonResponse({
            'status': 'error',
            'message': '잘못된 날짜 형식입니다.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'오류 주문 취소 중 문제가 발생했습니다: {str(e)}'
        }, status=500)
