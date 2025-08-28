from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.db.models import Sum, F, Count, Q
from datetime import datetime, date, timedelta
from django.db.models.functions import TruncDate
from django.db import transaction
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView
from django.contrib import messages
from .models import Center, Shipper, Courier, Product, User, Order, OrderItem, SalesChannel, StockMovement
from .forms import (
    CenterForm, ShipperForm, CourierForm, ProductForm, StockIOForm,
    StockUpdateForm, UserUpdateForm, CustomUserCreationForm, OrderUpdateForm
)
import openpyxl
from django.views.decorators.http import require_POST
import json


# ----------------------------------------
# 송장 출력 뷰
# ----------------------------------------
@login_required
@require_POST
def order_invoice_view(request):
    order_ids_str = request.POST.get('order_ids', '')
    if not order_ids_str:
        return HttpResponse("출력할 주문이 선택되지 않았습니다.", status=400)
    order_ids = [int(id) for id in order_ids_str.split(',')]
    orders = Order.objects.filter(id__in=order_ids).select_related('shipper__center').prefetch_related('items__product')
    if not orders:
        return HttpResponse("유효한 주문을 찾을 수 없습니다.", status=404)
    Order.objects.filter(id__in=order_ids).update(order_status='SHIPPED')
    context = {'orders': orders}
    return render(request, 'wms_app/invoice_template.html', context)

# ----------------------------------------
# API 뷰
# ----------------------------------------

@login_required
@require_POST
def check_duplicates_api(request):
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'error': '엑셀 파일이 없습니다.'}, status=400)
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        duplicate_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            shipper_name = str(row[1]).strip() if row[1] else None
            recipient_name = str(row[3]).strip() if row[3] else None
            recipient_phone = str(row[4]).strip() if row[4] else None
            address = str(row[5]).strip() if row[5] else None
            if not all([shipper_name, recipient_name, recipient_phone, address]):
                continue
            if Order.objects.filter(shipper__name=shipper_name, recipient_name=recipient_name, address=address, recipient_phone=recipient_phone).exists():
                duplicate_count += 1
        return JsonResponse({'has_duplicates': duplicate_count > 0, 'duplicate_count': duplicate_count})
    except Exception as e:
        return JsonResponse({'error': f'파일 검사 중 오류 발생: {str(e)}'}, status=500)

@login_required
@require_POST
@transaction.atomic
def process_orders_api(request):
    excel_file = request.FILES.get('excel_file')
    handle_duplicates = request.POST.get('handle_duplicates')
    if not excel_file:
        return JsonResponse({'status': 'error', 'message': '엑셀 파일이 없습니다.'}, status=400)

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active
    
    current_success_orders = []
    current_error_orders = []
    processed_in_this_file = set()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        order_data = {
            'order_no': str(row[0]).strip() if row[0] else None, 'shipper_name': str(row[1]).strip() if row[1] else None,
            'channel_name': str(row[2]).strip() if row[2] else None, 'recipient_name': str(row[3]).strip() if row[3] else None,
            'recipient_phone': str(row[4]).strip() if row[4] else '', 'address': str(row[5]).strip() if row[5] else '',
            'product_identifier': str(row[6]).strip() if row[6] else None, 'quantity': int(row[7]) if row[7] and str(row[7]).isdigit() else 0,
        }
        
        validation_errors = []
        error_fields = []
        shipper = None

        try:
            # 1. 모든 오류를 검출하여 리스트에 누적
            if not order_data['shipper_name']:
                validation_errors.append("화주사 정보 누락")
                error_fields.append('shipper_name')
            else:
                try:
                    shipper = Shipper.objects.get(name=order_data['shipper_name'])
                except Shipper.DoesNotExist:
                    validation_errors.append("미등록 화주사")
                    error_fields.append('shipper_name')

            if not order_data['product_identifier']:
                validation_errors.append("상품 정보 누락")
                error_fields.append('product_identifier')
            elif shipper:
                product_exists = Product.objects.filter(
                    Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), 
                    shipper=shipper
                ).exists()
                if not product_exists:
                    validation_errors.append("미등록 상품")
                    error_fields.append('product_identifier')

            if not order_data['quantity'] or order_data['quantity'] <= 0:
                validation_errors.append("수량 오류")
                error_fields.append('quantity')
            
            if not order_data['channel_name']:
                validation_errors.append("판매채널 누락")
                error_fields.append('channel_name')

            # 2. 누적된 오류가 있으면 예외 발생
            if validation_errors:
                raise ValueError(", ".join(sorted(list(set(validation_errors)))))

            # 3. 중복 검사
            is_duplicate_db = Order.objects.filter(
                shipper=shipper, recipient_name=order_data['recipient_name'], 
                address=order_data['address'], recipient_phone=order_data['recipient_phone']
            ).exists()
            if is_duplicate_db and handle_duplicates != 'yes':
                continue
            
            file_duplicate_key = (order_data['shipper_name'], order_data['recipient_name'], order_data['address'], order_data['recipient_phone'])
            if file_duplicate_key in processed_in_this_file and handle_duplicates != 'yes':
                continue
            processed_in_this_file.add(file_duplicate_key)

            # 4. 성공 처리
            channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
            product = Product.objects.get(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper)
            
            if handle_duplicates == 'yes' and is_duplicate_db:
                order_data['order_no'] = None
                
            order = Order.objects.create(
                order_no=order_data['order_no'], shipper=shipper, channel=channel,
                recipient_name=order_data['recipient_name'], recipient_phone=order_data['recipient_phone'],
                address=order_data['address'], order_date=datetime.now(), order_status='PENDING'
            )
            OrderItem.objects.create(order=order, product=product, quantity=order_data['quantity'])
            current_success_orders.append(order)

        except Exception as e:
            # 5. 오류 기록
            error_data_packet = {
                'row_idx': row_idx,
                'error_message': str(e),
                'error_fields': sorted(list(set(error_fields))),
                'original_data': order_data,
            }
            try:
                channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
                error_order = Order.objects.create(
                    order_no=order_data['order_no'] or f"NO_ID_ERROR-{row_idx}",
                    shipper=shipper, channel=channel,
                    recipient_name=order_data['recipient_name'], recipient_phone=order_data['recipient_phone'],
                    address=order_data['address'], order_date=datetime.now(), 
                    order_status='ERROR', error_message=json.dumps(error_data_packet)
                )
                current_error_orders.append(error_order)
            except Exception:
                current_error_orders.append(error_data_packet)

    messages.success(request, f"엑셀 처리가 완료되었습니다. 성공: {len(current_success_orders)}건, 실패: {len(current_error_orders)}건")
    request.session['temp_errors'] = [e for e in current_error_orders if isinstance(e, dict)]
    today_str = date.today().strftime('%Y-%m-%d')
    return JsonResponse({'status': 'success', 'redirect_url': reverse('order_manage') + f'?date={today_str}'})

@login_required
@require_POST
@transaction.atomic
def batch_retry_error_api(request):
    all_items_data = json.loads(request.body)
    results = []
    for item_data in all_items_data:
        order_data = item_data.get('data', {})
        unique_id = item_data.get('unique_id')
        
        validation_errors = []
        error_fields = []
        shipper = None
        
        try:
            if not order_data.get('shipper_name'):
                validation_errors.append("화주사 정보 누락")
                error_fields.append('shipper_name')
            else:
                try:
                    shipper = Shipper.objects.get(name=order_data['shipper_name'])
                except Shipper.DoesNotExist:
                    validation_errors.append("미등록 화주사")
                    error_fields.append('shipper_name')

            if not order_data.get('product_identifier'):
                validation_errors.append("상품 정보 누락")
                error_fields.append('product_identifier')
            elif shipper:
                product_exists = Product.objects.filter(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper).exists()
                if not product_exists:
                    validation_errors.append("미등록 상품")
                    error_fields.append('product_identifier')

            if not order_data.get('quantity') or int(order_data.get('quantity', 0)) <= 0:
                validation_errors.append("수량 오류")
                error_fields.append('quantity')

            if shipper and not error_fields:
                is_duplicate = Order.objects.filter(
                    shipper=shipper, recipient_name=order_data['recipient_name'], 
                    address=order_data['address'], recipient_phone=order_data['recipient_phone'],
                    order_status__in=['PENDING', 'PROCESSING', 'SHIPPED', 'DELIVERED']
                ).exists()
                if is_duplicate:
                    validation_errors.append("중복 주문")
                    error_fields.append('multiple')

            if validation_errors:
                raise ValueError(", ".join(sorted(list(set(validation_errors)))))
            
            channel, _ = SalesChannel.objects.get_or_create(name=order_data['channel_name'])
            product = Product.objects.get(Q(barcode=order_data['product_identifier']) | Q(name=order_data['product_identifier']), shipper=shipper)
            
            new_order = Order.objects.create(
                order_no=order_data.get('order_no') or None, shipper=shipper, channel=channel,
                recipient_name=order_data['recipient_name'], recipient_phone=order_data['recipient_phone'],
                address=order_data['address'], order_date=datetime.now(), order_status='PENDING'
            )
            OrderItem.objects.create(order=new_order, product=product, quantity=order_data['quantity'])
            
            id_type, id_value = unique_id.split('-')
            if id_type == 'db':
                Order.objects.filter(id=id_value).delete()
            elif id_type == 'session' and 'temp_errors' in request.session:
                request.session['temp_errors'] = [e for e in request.session['temp_errors'] if e.get('row_idx') != int(id_value)]
                request.session.modified = True
            
            results.append({'unique_id': unique_id, 'status': 'success'})

        except Exception as e:
            results.append({
                'unique_id': unique_id, 'status': 'error',
                'error_message': str(e),
                'error_fields': sorted(list(set(error_fields))),
            })
            
    return JsonResponse({'results': results})

def translate_error_message(message_str):
    return message_str

@login_required
def channel_order_chart_data(request):
    date_str = request.GET.get('date')
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    orders_for_date = Order.objects.filter(order_date__date=target_date)
    channel_counts = SalesChannel.objects.filter(order__in=orders_for_date).annotate(order_count=Count('order')).values('name', 'order_count')
    labels = [data['name'] for data in channel_counts]
    data = [data['order_count'] for data in channel_counts]
    return JsonResponse({'labels': labels, 'data': data})

@login_required
def order_chart_data(request):
    start_date_str = request.GET.get('start')
    end_date_str = request.GET.get('end')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        end_date = date.today()
        start_date = end_date - timedelta(days=6)
    labels = []
    current_date = start_date
    while current_date <= end_date:
        labels.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    status_map = { 'PENDING': '주문접수', 'PROCESSING': '처리중', 'ERROR': '오류' }
    orders = Order.objects.filter(
            order_date__date__range=[start_date, end_date],
            order_status__in=status_map.keys()
        ).annotate(date=TruncDate('order_date')) \
         .values('date', 'order_status') \
         .annotate(count=Count('id')) \
         .order_by('date')
    daily_counts = {label: {status: 0 for status in status_map} for label in labels}
    for order in orders:
        date_str = order['date'].strftime('%Y-%m-%d')
        if date_str in daily_counts:
            daily_counts[date_str][order['order_status']] = order['count']
    status_colors = { 'PENDING': 'rgba(54, 162, 235, 0.7)', 'PROCESSING': 'rgba(255, 159, 64, 0.7)', 'ERROR': 'rgba(255, 99, 132, 0.7)' }
    final_datasets = []
    for status_code, status_name in status_map.items():
        data = [daily_counts[label][status_code] for label in labels]
        if any(d > 0 for d in data):
            final_datasets.append({
                'label': status_name, 'data': data,
                'borderColor': status_colors.get(status_code, 'rgba(0, 0, 0, 0.7)'),
                'backgroundColor': status_colors.get(status_code, 'rgba(0, 0, 0, 0.7)'),
                'fill': False, 'tension': 0.1
            })
    return JsonResponse({'labels': labels, 'datasets': final_datasets})

def check_username(request):
    username = request.GET.get('username', None)
    is_taken = User.objects.filter(username__iexact=username).exists()
    data = {'is_available': not is_taken}
    return JsonResponse(data)

def wms_logout_view(request):
    logout(request)
    return redirect('login')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    def form_invalid(self, form):
        messages.error(self.request, '아이디 또는 비밀번호가 올바르지 않습니다.')
        return super().form_invalid(form)
    def form_valid(self, form):
        user = form.get_user()
        if not user.is_active:
            messages.error(self.request, '아직 승인되지 않은 계정입니다. 관리자에게 문의하세요.')
            return self.form_invalid(form)
        return super().form_valid(form)

def signup_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('signup_done')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

def signup_done_view(request):
    return render(request, 'registration/signup_done.html')

@login_required
def dashboard(request):
    context = {'page_title': '홈', 'active_menu': 'dashboard'}
    return render(request, 'wms_app/order_list_page.html', context)

@login_required
def order_manage(request):
    date_str = request.GET.get('date')
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else date.today()
    orders_for_date = Order.objects.filter(order_date__date=selected_date)
    total_count = orders_for_date.count()
    success_count = orders_for_date.exclude(order_status='ERROR').count()
    error_count_db = orders_for_date.filter(order_status='ERROR').count()
    temp_errors_count = len(request.session.get('temp_errors', []))
    total_count += temp_errors_count
    error_count = error_count_db + temp_errors_count
    daily_stats = {
        'total_count': total_count, 'success_count': success_count, 'error_count': error_count,
        'success_rate': (success_count / total_count * 100) if total_count > 0 else 0,
        'error_rate': (error_count / total_count * 100) if total_count > 0 else 0,
    }
    context = {
        'page_title': '주문 관리', 'active_menu': 'orders',
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'daily_stats': daily_stats,
    }
    return render(request, 'wms_app/order_manage.html', context)

@login_required
def order_list_success(request, date_str):
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    orders = Order.objects.filter(order_date__date=target_date).exclude(order_status='ERROR').prefetch_related('items__product')
    orders_json = []
    for order in orders:
        items = [{'product_name': item.product.name, 'quantity': item.quantity} for item in order.items.all()]
        orders_json.append({'id': order.id, 'items': items})
    context = {
        'page_title': f'{date_str} 성공 주문 목록', 'orders': orders,
        'list_type': 'success', 'orders_json': orders_json,
    }
    return render(request, 'wms_app/order_list_result.html', context)

@login_required
def order_list_error(request, date_str):
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    db_errors = []
    for order in Order.objects.filter(order_date__date=target_date, order_status='ERROR'):
        try:
            error_details = json.loads(order.error_message)
            order.original_data = error_details.get('original_data', {})
            order.error_message_translated = error_details.get('error_message', '')
            order.error_fields = error_details.get('error_fields', []) 
        except (json.JSONDecodeError, TypeError):
            order.original_data = {}
            order.error_message_translated = order.error_message
            order.error_fields = []
        db_errors.append(order)
        
    temp_errors = request.session.pop('temp_errors', [])
    for error in temp_errors:
        error['error_message_translated'] = error.get('error_message', '')
        error['error_fields'] = error.get('error_fields', [])

    all_errors = db_errors + temp_errors
    
    orders_json = []
    for order in db_errors:
        if hasattr(order, 'items') and order.items.exists():
            items = [{'product_name': item.product.name, 'quantity': item.quantity} for item in order.items.all()]
            orders_json.append({'id': order.id, 'items': items})

    context = {
        'page_title': f'{date_str} 오류 주문 목록', 'orders': all_errors,
        'list_type': 'error', 'orders_json': orders_json,
    }
    return render(request, 'wms_app/order_list_result.html', context)

@login_required
def download_error_excel(request, date_str):
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponseBadRequest("잘못된 날짜 형식입니다. YYYY-MM-DD 형식으로 입력해주세요.")
    error_orders = Order.objects.filter(order_date__date=target_date, order_status='ERROR')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "오류 주문 목록"
    headers = ["주문번호", "화주사", "판매채널", "수취인명", "연락처", "주소", "상품 식별자", "수량", "오류 원인"]
    ws.append(headers)
    for order in error_orders:
        original_data = {}
        error_message = order.error_message
        try:
            error_details = json.loads(order.error_message)
            original_data = error_details.get('original_data', {})
            error_message = error_details.get('error_message', error_message)
        except (json.JSONDecodeError, TypeError):
            original_data['order_no'] = order.order_no
            original_data['shipper_name'] = order.shipper.name if order.shipper else ''
            original_data['channel_name'] = order.channel.name if order.channel else ''
            original_data['recipient_name'] = order.recipient_name
            original_data['recipient_phone'] = order.recipient_phone
            original_data['address'] = order.address
        row_data = [
            original_data.get('order_no', ''), original_data.get('shipper_name', ''),
            original_data.get('channel_name', ''), original_data.get('recipient_name', ''),
            original_data.get('recipient_phone', ''), original_data.get('address', ''),
            original_data.get('product_identifier', ''), original_data.get('quantity', ''),
            translate_error_message(error_message)
        ]
        ws.append(row_data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="error_orders_{date_str}.xlsx"'
    wb.save(response)
    return response

@login_required
def order_update_view(request, order_pk):
    order = get_object_or_404(Order, pk=order_pk)
    if request.method == 'POST':
        form = OrderUpdateForm(request.POST, instance=order)
        if form.is_valid():
            updated_order = form.save(commit=False)
            updated_order.error_message = ""
            updated_order.order_status = 'PENDING'
            updated_order.save()
            messages.success(request, f"주문({order.order_no})이 성공적으로 수정되었습니다.")
            date_str = order.order_date.strftime('%Y-%m-%d')
            return redirect('order_list_error', date_str=date_str)
    else:
        try:
            error_details = json.loads(order.error_message)
            initial_data = error_details.get('original_data', {})
            initial_data.update({
                'shipper': order.shipper, 'channel': order.channel, 'order_no': order.order_no,
                'recipient_name': order.recipient_name, 'recipient_phone': order.recipient_phone,
                'address': order.address,
            })
            form = OrderUpdateForm(initial=initial_data)
        except (json.JSONDecodeError, TypeError):
            form = OrderUpdateForm(instance=order)
    context = { 'page_title': '오류 주문 수정', 'form': form, 'order': order }
    return render(request, 'wms_app/order_update_form.html', context)

@login_required
def stock_in(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '입고', 'active_menu': 'inout'})

@login_required
def stock_out(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '출고', 'active_menu': 'inout'})

@login_required
def settlement_status(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '정산 현황', 'active_menu': 'settlement'})

@login_required
def settlement_billing(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '정산 청구내역', 'active_menu': 'settlement'})

@login_required
def settlement_config(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '정산내역설정', 'active_menu': 'settlement'})

@login_required
def order_manage_new(request):
    return render(request, 'wms_app/placeholder_page.html', {'page_title': '주문 관리', 'active_menu': 'management'})
    
@login_required
def management_dashboard(request):
    context = {
        'page_title': '통합 관리', 'active_menu': 'management',
        'shipper_count': Shipper.objects.count(), 'product_count': Product.objects.count(),
        'center_count': Center.objects.count(), 'courier_count': Courier.objects.count(),
    }
    return render(request, 'wms_app/management_dashboard.html', context)

@login_required
def stock_manage(request):
    queryset = Product.objects.select_related('shipper__center').all()
    selected_center = request.session.get('selected_center')
    selected_shipper = request.session.get('selected_shipper')
    if selected_center:
        queryset = queryset.filter(shipper__center__name=selected_center)
    if selected_shipper:
        queryset = queryset.filter(shipper__name=selected_shipper)
    context = {
        'page_title': '재고관리', 'object_list': queryset,
        'columns': [
            {'header': '상품명', 'key': 'name'}, {'header': '크기(cm)', 'is_size': True},
            {'header': '재고', 'key': 'quantity'}, {'header': '바코드', 'key': 'barcode'},
            {'header': '화주사명', 'key': 'shipper'},
        ], 'active_menu': 'management', 'update_url_name': 'stock_update',
    }
    return render(request, 'wms_app/generic_list.html', context)

@login_required
@transaction.atomic
def stock_io_view(request):
    if request.method == 'POST':
        form_data = request.POST.copy()
        form_data['product'] = request.POST.get('product')
        form = StockIOForm(form_data)
        if form.is_valid():
            product = form.cleaned_data['product']
            quantity = form.cleaned_data['quantity']
            memo = form.cleaned_data['memo']
            io_type = request.POST.get('io_type')
            if io_type == 'in':
                product.quantity = F('quantity') + quantity
                movement_type = 'IN'
            elif io_type == 'out':
                product.refresh_from_db()
                if product.quantity < quantity:
                    return HttpResponseBadRequest("재고가 부족합니다.")
                product.quantity = F('quantity') - quantity
                movement_type = 'OUT'
            product.save()
            StockMovement.objects.create(product=product, movement_type=movement_type, quantity=quantity, memo=memo)
            return redirect('stock_io')
    products = Product.objects.select_related('shipper__center').all()
    selected_center = request.session.get('selected_center')
    selected_shipper = request.session.get('selected_shipper')
    if selected_center:
        products = products.filter(shipper__center__name=selected_center)
    if selected_shipper:
        products = products.filter(shipper__name=selected_shipper)
    context = {'page_title': '재고 입출고', 'products': products, 'active_menu': 'inout'}
    return render(request, 'wms_app/stock_io.html', context)

@login_required
def stock_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = StockUpdateForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('stock_manage')
    else:
        form = StockUpdateForm(instance=product)
    context = { 'form': form, 'product': product, 'page_title': '재고 수량 수정', 'active_menu': 'management' }
    return render(request, 'wms_app/stock_update_form.html', context)

@login_required
def stock_movement_history(request):
    movements = StockMovement.objects.select_related('product__shipper').order_by('-timestamp')
    context = { 'page_title': '입출고 기록', 'movements': movements, 'active_menu': 'inout' }
    return render(request, 'wms_app/stock_history.html', context)

@login_required
def user_manage(request):
    user_list = User.objects.filter(is_superuser=False)
    context = { 'page_title': '사용자 관리', 'active_menu': 'management', 'user_list': user_list, }
    return render(request, 'wms_app/user_list.html', context)

@login_required
def user_update(request, pk):
    user_instance = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=user_instance)
        if form.is_valid():
            form.save()
            return redirect('user_manage')
    else:
        form = UserUpdateForm(instance=user_instance)
    context = { 'form': form, 'target_user': user_instance, 'page_title': '사용자 역할 및 소속 수정', 'active_menu': 'management' }
    return render(request, 'wms_app/user_form.html', context)

class CenterListView(LoginRequiredMixin, ListView):
    model = Center
    template_name = 'wms_app/generic_list.html'
    context_object_name = 'object_list'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': '센터 관리', 'columns': [{'header': '센터명', 'key': 'name'}, {'header': '주소', 'key': 'address'}],
            'create_url': reverse_lazy('center_create'), 'update_url_name': 'center_update', 'delete_url_name': 'center_delete',
            'active_menu': 'management'
        })
        return context

class CenterCreateView(LoginRequiredMixin, CreateView):
    model = Center
    form_class = CenterForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('center_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '센터 등록'
        context['active_menu'] = 'management'
        return context

class CenterUpdateView(LoginRequiredMixin, UpdateView):
    model = Center
    form_class = CenterForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('center_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '센터 편집'
        context['active_menu'] = 'management'
        return context

class CenterDeleteView(LoginRequiredMixin, DeleteView):
    model = Center
    success_url = reverse_lazy('center_list')

class ShipperListView(LoginRequiredMixin, ListView):
    model = Shipper
    template_name = 'wms_app/generic_list.html'
    context_object_name = 'object_list'
    def get_queryset(self):
        queryset = super().get_queryset().select_related('center')
        selected_center = self.request.session.get('selected_center')
        if selected_center:
            queryset = queryset.filter(center__name=selected_center)
        return queryset
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': '화주사 관리', 'columns': [{'header': '화주사명', 'key': 'name'}, {'header': '담당자', 'key': 'contact'}, {'header': '소속 센터', 'key': 'center'}],
            'create_url': reverse_lazy('shipper_create'), 'update_url_name': 'shipper_update', 'delete_url_name': 'shipper_delete',
            'extra_actions': [{'label': '판매 상품', 'url_name': 'shipper_product_list', 'class': 'btn-info'}], 'active_menu': 'management'
        })
        return context

class ShipperCreateView(LoginRequiredMixin, CreateView):
    model = Shipper
    form_class = ShipperForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('shipper_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '화주사 등록'
        context['active_menu'] = 'management'
        return context

class ShipperUpdateView(LoginRequiredMixin, UpdateView):
    model = Shipper
    form_class = ShipperForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('shipper_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '화주사 편집'
        context['active_menu'] = 'management'
        return context

class ShipperDeleteView(LoginRequiredMixin, DeleteView):
    model = Shipper
    success_url = reverse_lazy('shipper_list')

class CourierListView(LoginRequiredMixin, ListView):
    model = Courier
    template_name = 'wms_app/generic_list.html'
    context_object_name = 'object_list'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': '택배사 관리', 'columns': [{'header': '택배사명', 'key': 'name'}, {'header': '연락처', 'key': 'contact'}],
            'create_url': reverse_lazy('courier_create'), 'update_url_name': 'courier_update', 'delete_url_name': 'courier_delete',
            'active_menu': 'management'
        })
        return context

class CourierCreateView(LoginRequiredMixin, CreateView):
    model = Courier
    form_class = CourierForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('courier_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '택배사 등록'
        context['active_menu'] = 'management'
        return context

class CourierUpdateView(LoginRequiredMixin, UpdateView):
    model = Courier
    form_class = CourierForm
    template_name = 'wms_app/generic_form.html'
    success_url = reverse_lazy('courier_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = '택배사 편집'
        context['active_menu'] = 'management'
        return context

class CourierDeleteView(LoginRequiredMixin, DeleteView):
    model = Courier
    success_url = reverse_lazy('courier_list')

@login_required
def shipper_product_list(request, shipper_pk):
    shipper = get_object_or_404(Shipper, pk=shipper_pk)
    products = Product.objects.filter(shipper=shipper)
    context = { 'shipper': shipper, 'products': products, 'page_title': f'{shipper.name} 판매 상품', 'active_menu': 'management' }
    return render(request, 'wms_app/shipper_product_list.html', context)

@login_required
def shipper_product_create(request, shipper_pk):
    shipper = get_object_or_404(Shipper, pk=shipper_pk)
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            product.shipper = shipper
            product.save()
            return redirect('shipper_product_list', shipper_pk=shipper.pk)
    else:
        form = ProductForm()
    context = { 'form': form, 'shipper': shipper, 'page_title': f'{shipper.name} 상품 등록', 'active_menu': 'management' }
    return render(request, 'wms_app/shipper_product_form.html', context)

@login_required
def shipper_product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('shipper_product_list', shipper_pk=product.shipper.pk)
    else:
        form = ProductForm(instance=product)
    context = { 'form': form, 'page_title': '판매 상품 편집', 'active_menu': 'management' }
    return render(request, 'wms_app/shipper_product_form.html', context)

@login_required
def shipper_product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    shipper_pk = product.shipper.pk
    if request.method == 'POST':
        product.delete()
        return redirect('shipper_product_list', shipper_pk=shipper_pk)
    return HttpResponseBadRequest("잘못된 요청입니다.")

def filters(request):
    selected_center_name = request.session.get('selected_center', '')
    shippers = Shipper.objects.all()
    if selected_center_name:
        shippers = shippers.filter(center__name=selected_center_name)
    return {
        'centers': Center.objects.all(), 'shippers': shippers,
        'selected_center': selected_center_name, 'selected_shipper': request.session.get('selected_shipper', ''),
    }