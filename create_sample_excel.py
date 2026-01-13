import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "주문업로드_예시"

# 헤더 설정
headers = ['주문번호', '화주사', '판매채널', '수취인', '연락처', '주소', '상품명', '수량', '배송메모']
ws.append(headers)

# 스타일 적용
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# 예시 데이터 (성공 케이스 + 실패 케이스 섞음)
data = [
    # 1. 정상 데이터
    ['ORD-20240108-01', '테스트화주', '스마트스토어', '김철수', '010-1234-5678', '서울시 강남구 역삼동 123-45', '테스트상품A', 2, '안전배송'],
    ['ORD-20240108-02', '테스트화주', '쿠팡', '이영희', '010-9876-5432', '경기도 성남시 분당구 판교동 55', '테스트상품B', 1, '부재시 문앞'],
    
    # 2. 오류 데이터 (화주사 누락) - 직접 수정 테스트용
    ['ORD-20240108-03', '', '11번가', '박민수', '010-1111-2222', '부산시 해운대구 우동 77', '테스트상품A', 3, '빠른배송'],
    
    # 3. 오류 데이터 (상품명 불일치) - 자동완성 테스트용
    ['ORD-20240108-04', '테스트화주', 'G마켓', '최지우', '010-3333-4444', '대구시 수성구 범어동 88', '없는상품X', 1, ''],
    
    # 4. 오류 데이터 (수량 오류)
    ['ORD-20240108-05', '테스트화주', '자사몰', '정우성', '010-5555-6666', '광주시 서구 치평동 99', '테스트상품B', 0, ''],
]

for row in data:
    ws.append(row)

# 열 너비 조정
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 20

# 저장
filename = r"c:\Users\one09\Desktop\wms\sample_orders.xlsx"
wb.save(filename)
print(f"File created at: {filename}")
